#!/usr/bin/env python3
"""
ImpactEng — Agentic AI Readiness Audit Scorecard Generator
Produces: ImpactEng_AI_Readiness_Audit_Scorecard.xlsx
Requires: pip install openpyxl>=3.1
Run:      python3 scripts/generate_scorecard.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

OUTPUT_FILE = "ImpactEng_AI_Readiness_Audit_Scorecard.xlsx"

# ── Brand colours ──────────────────────────────────────────────────────────────
NAVY   = "1E3A5F"
ACCENT = "2E86AB"
LTBLUE = "D6EAF8"
WHITE  = "FFFFFF"
AMBER  = "FFC000"
RED    = "FF6B6B"
GREEN  = "70AD47"
GREY   = "F2F2F2"

# ── Assessment dimensions ──────────────────────────────────────────────────────
DIMENSIONS = [
    {
        "name": "Cloud Infrastructure Readiness",
        "weight": 0.20,
        "questions": [
            "Cloud architecture is designed for variable, high-throughput workloads (not just static apps)",
            "Observability and monitoring tooling is in place (logs, metrics, traces)",
            "Cost controls and budget alerting are configured",
            "Auto-scaling and reliability patterns (circuit breakers, retries) are implemented",
            "Security posture: IAM, secrets management, and network controls are mature",
        ],
    },
    {
        "name": "Data Quality & Pipeline Maturity",
        "weight": 0.20,
        "questions": [
            "Data is accessible via APIs or queryable data stores (not locked in spreadsheets/silos)",
            "Data quality is actively monitored — completeness, accuracy, freshness",
            "Data pipelines are reliable, tested, and have alerting on failures",
            "A data governance policy exists and is followed",
            "Sensitive/PII data is classified and controls are in place",
        ],
    },
    {
        "name": "DevOps & MLOps Practices",
        "weight": 0.20,
        "questions": [
            "CI/CD pipelines are automated and used for all production deployments",
            "Deployment frequency is weekly or better; rollback is automated",
            "There is a defined process for deploying and versioning AI/ML models",
            "Incident response runbooks exist and are tested",
            "Testing coverage includes integration and regression tests for critical paths",
        ],
    },
    {
        "name": "Team & Organisational Capability",
        "weight": 0.15,
        "questions": [
            "At least one team member has hands-on experience deploying AI/ML in production",
            "There is a named owner for the AI initiative (not just a committee)",
            "The team has capacity allocated (not just 'when we have time')",
            "Leadership understands AI limitations and failure modes (not just the hype)",
            "Cross-functional alignment exists between engineering, data, and business stakeholders",
        ],
    },
    {
        "name": "Governance & Responsible AI",
        "weight": 0.15,
        "questions": [
            "A policy exists for what data AI models can and cannot access",
            "Human-in-the-loop checkpoints are defined for high-risk AI decisions",
            "Audit logging is in place for AI-influenced decisions",
            "A process exists to detect and respond to AI model drift or degraded performance",
            "The organisation is aware of and planning for relevant AI regulations (AU AI Ethics Framework, EU AI Act if applicable)",
        ],
    },
    {
        "name": "Strategic Alignment & Use Case Prioritisation",
        "weight": 0.10,
        "questions": [
            "There are 1-3 specific, named AI use cases with defined success metrics",
            "The use cases are prioritised by business value, not just technical interest",
            "A realistic ROI estimate exists for the top use case",
            "The board or senior leadership has explicitly endorsed the AI program",
            "There is a defined timeline with milestones (not an open-ended exploration)",
        ],
    },
]

assert abs(sum(d["weight"] for d in DIMENSIONS) - 1.0) < 1e-9, "Weights must sum to 1.0"


# ── Style helpers ──────────────────────────────────────────────────────────────

def solid_fill(hex_colour):
    return PatternFill(fill_type="solid", fgColor=hex_colour)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ── Sheet 1: Client Info ───────────────────────────────────────────────────────

def build_client_info(wb):
    ws = wb.create_sheet("Client Info")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "ImpactEng — Agentic AI Readiness Audit"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells("A1:C1")

    ws["A2"] = "Client Information"
    ws["A2"].font = Font(bold=True, size=11, color=ACCENT)
    ws.merge_cells("A2:C2")

    fields = [
        ("Client Name", ""),
        ("Engagement Date", ""),
        ("Tier", "Starter / Professional / Enterprise"),
        ("Lead Consultant", ""),
        ("Cloud Provider", "AWS / Azure / GCP / Other"),
        ("Team Size", ""),
        ("Industry", ""),
    ]

    for i, (label, hint) in enumerate(fields, start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True, size=11)
        label_cell.fill = solid_fill(LTBLUE)
        label_cell.border = thin_border()

        hint_cell = ws.cell(row=i, column=2, value=hint)
        hint_cell.font = Font(italic=True, color="888888")
        hint_cell.border = thin_border()

        ws.cell(row=i, column=3).border = thin_border()

    set_width(ws, "A", 28)
    set_width(ws, "B", 42)
    set_width(ws, "C", 16)


# ── Sheet 2: Scoring ───────────────────────────────────────────────────────────

def build_scoring(wb):
    ws = wb.create_sheet("Scoring")
    ws.sheet_view.showGridLines = False

    # Column headers
    col_defs = [
        ("Question", 62),
        ("Score (1-5)", 14),
        ("Evidence / Notes", 42),
        ("Priority", 12),
    ]
    for col_idx, (header, width) in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = solid_fill(NAVY)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()
        set_width(ws, get_column_letter(col_idx), width)

    # Score (1-5) dropdown
    dv_score = DataValidation(
        type="list",
        formula1='"1,2,3,4,5"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid score",
        error="Please enter a value between 1 and 5.",
    )
    ws.add_data_validation(dv_score)

    # Priority dropdown
    dv_priority = DataValidation(
        type="list",
        formula1='"High,Med,Low"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_priority)

    current_row = 2
    section_meta = []

    for dim in DIMENSIONS:
        # Section header row
        header_cell = ws.cell(row=current_row, column=1, value=dim["name"])
        header_cell.font = Font(bold=True, size=11, color=WHITE)
        header_cell.fill = solid_fill(ACCENT)
        header_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=4,
        )
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        q_start_row = current_row

        for q_text in dim["questions"]:
            ws.cell(row=current_row, column=1, value=q_text).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=current_row, column=1).border = thin_border()
            ws.row_dimensions[current_row].height = 32

            score_cell = ws.cell(row=current_row, column=2)
            score_cell.alignment = Alignment(horizontal="center", vertical="top")
            score_cell.border = thin_border()
            dv_score.add(score_cell.coordinate)

            notes_cell = ws.cell(row=current_row, column=3)
            notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
            notes_cell.border = thin_border()

            priority_cell = ws.cell(row=current_row, column=4)
            priority_cell.alignment = Alignment(horizontal="center", vertical="top")
            priority_cell.border = thin_border()
            dv_priority.add(priority_cell.coordinate)

            current_row += 1

        q_end_row = current_row - 1

        # Section average row
        avg_row = current_row
        ws.cell(row=avg_row, column=1, value="Section Average").font = Font(bold=True, size=11)
        ws.cell(row=avg_row, column=1).fill = solid_fill(GREY)
        ws.cell(row=avg_row, column=1).border = thin_border()

        avg_cell = ws.cell(
            row=avg_row, column=2,
            value=f"=IFERROR(AVERAGE(B{q_start_row}:B{q_end_row}),\"\")"
        )
        avg_cell.number_format = "0.00"
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.fill = solid_fill(GREY)
        avg_cell.font = Font(bold=True, size=11)
        avg_cell.border = thin_border()

        ws.cell(row=avg_row, column=3, value=f"Weight: {int(dim['weight'] * 100)}%").font = Font(italic=True)
        ws.cell(row=avg_row, column=3).border = thin_border()
        ws.cell(row=avg_row, column=4).border = thin_border()
        current_row += 1

        # Weighted score row
        weighted_row = current_row
        ws.cell(row=weighted_row, column=1, value="Weighted Score").font = Font(bold=True, size=11)
        ws.cell(row=weighted_row, column=1).fill = solid_fill(LTBLUE)
        ws.cell(row=weighted_row, column=1).border = thin_border()

        weighted_cell = ws.cell(
            row=weighted_row, column=2,
            value=f"=IFERROR(B{avg_row}*{dim['weight']},\"\")"
        )
        weighted_cell.number_format = "0.00"
        weighted_cell.alignment = Alignment(horizontal="center")
        weighted_cell.fill = solid_fill(LTBLUE)
        weighted_cell.font = Font(bold=True, size=11)
        weighted_cell.border = thin_border()

        ws.cell(row=weighted_row, column=3).border = thin_border()
        ws.cell(row=weighted_row, column=4).border = thin_border()
        current_row += 1

        section_meta.append({
            "name": dim["name"],
            "weight": dim["weight"],
            "avg_row": avg_row,
            "weighted_row": weighted_row,
        })

        # Blank separator row
        current_row += 1

    # Conditional formatting on score column
    score_range = f"B2:B{current_row}"
    red_fill   = PatternFill(fill_type="solid", fgColor=RED)
    amber_fill = PatternFill(fill_type="solid", fgColor=AMBER)
    green_fill = PatternFill(fill_type="solid", fgColor=GREEN)

    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="between", formula=["1", "2"], fill=red_fill)
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="equal", formula=["3"], fill=amber_fill)
    )
    ws.conditional_formatting.add(
        score_range,
        CellIsRule(operator="between", formula=["4", "5"], fill=green_fill)
    )

    return ws, section_meta


# ── Sheet 3: Results Dashboard ─────────────────────────────────────────────────

def build_dashboard(wb, section_meta):
    ws = wb.create_sheet("Results Dashboard")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Agentic AI Readiness — Results Dashboard"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells("A1:F1")

    # Overall score
    weighted_refs = "+".join(f"Scoring!B{m['weighted_row']}" for m in section_meta)
    ws["A3"] = "Overall Readiness Score (0–5)"
    ws["A3"].font = Font(bold=True, size=12)
    ws["B3"] = f"=IFERROR({weighted_refs},\"\")"
    ws["B3"].number_format = "0.00"
    ws["B3"].font = Font(bold=True, size=14, color=ACCENT)
    ws["B3"].alignment = Alignment(horizontal="center")

    # Readiness label
    ws["A4"] = "Readiness Level"
    ws["A4"].font = Font(bold=True, size=11)
    ws["B4"] = (
        '=IF(B3="","Score all dimensions first",'
        'IF(B3<2.5,"Not Ready \u2014 Foundational Work Required",'
        'IF(B3<3.5,"Partially Ready \u2014 Targeted Improvements Needed",'
        'IF(B3<4.5,"Ready \u2014 Accelerate with Confidence",'
        '"Leading \u2014 Scale and Govern"))))'
    )
    ws["B4"].font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells("B4:F4")

    # Per-dimension table (chart data source)
    headers = ["Dimension", "Weighted Score", "Max Possible"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = solid_fill(NAVY)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()

    for i, m in enumerate(section_meta, start=7):
        ws.cell(row=i, column=1, value=m["name"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=1).border = thin_border()

        score_cell = ws.cell(row=i, column=2, value=f"=Scoring!B{m['weighted_row']}")
        score_cell.number_format = "0.00"
        score_cell.alignment = Alignment(horizontal="center")
        score_cell.border = thin_border()

        max_cell = ws.cell(row=i, column=3, value=m["weight"])
        max_cell.number_format = "0.00"
        max_cell.alignment = Alignment(horizontal="center")
        max_cell.border = thin_border()

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Readiness Score by Dimension"
    chart.y_axis.title = "Weighted Score"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=6, max_row=6 + len(section_meta))
    cats_ref = Reference(ws, min_col=1, min_row=7, max_row=6 + len(section_meta))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "E6")

    # Recommended next step
    step_row = 7 + len(section_meta) + 2
    ws.cell(row=step_row, column=1, value="Recommended Next Step").font = Font(bold=True, size=12)
    ws.cell(row=step_row + 1, column=1, value=(
        '=IF(B3="","Score all dimensions first",'
        'IF(B3<2.5,"Path A: Cloud & data foundation work required before any AI deployment.",'
        'IF(B3<3.5,"Path B: Targeted improvements in lowest-scoring dimensions, then pilot.",'
        'IF(B3<4.5,"Path C: Proceed with a controlled AI agent pilot; governance review.",'
        '"Path D: Scale AI agents with continuous governance and optimisation."))))'
    ))
    ws.cell(row=step_row + 1, column=1).font = Font(bold=True, size=11, color=ACCENT)
    ws.merge_cells(start_row=step_row + 1, start_column=1, end_row=step_row + 1, end_column=5)

    set_width(ws, "A", 46)
    set_width(ws, "B", 18)
    set_width(ws, "C", 14)


# ── Sheet 4: Roadmap Template ──────────────────────────────────────────────────

def build_roadmap(wb):
    ws = wb.create_sheet("Roadmap Template")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "90-Day AI Readiness Roadmap"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells("A1:G1")

    ws["A2"] = "Complete column B (Gap Identified) from your Scoring sheet after the assessment readout."
    ws["A2"].font = Font(italic=True, color="888888")
    ws.merge_cells("A2:G2")

    col_defs = [
        ("Priority", 10),
        ("Gap Identified", 36),
        ("Recommended Action", 42),
        ("Owner", 18),
        ("Sprint", 24),
        ("Effort (S/M/L)", 16),
        ("Status", 14),
    ]
    for col_idx, (header, width) in enumerate(col_defs, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color=WHITE)
        cell.fill = solid_fill(NAVY)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()
        set_width(ws, get_column_letter(col_idx), width)

    # 15 pre-populated rows
    tiers = (
        [("High", RED,   "Sprint 1 (Days 1\u201330)")] * 5 +
        [("Med",  AMBER, "Sprint 2 (Days 31\u201360)")] * 5 +
        [("Low",  GREEN, "Sprint 3 (Days 61\u201390)")] * 5
    )

    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Blocked"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_status)

    dv_effort = DataValidation(
        type="list",
        formula1='"S,M,L"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_effort)

    for i, (priority, colour, sprint) in enumerate(tiers, start=4):
        priority_cell = ws.cell(row=i, column=1, value=priority)
        priority_cell.fill = solid_fill(colour)
        priority_cell.font = Font(bold=True, color=WHITE)
        priority_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=5, value=sprint)
        ws.cell(row=i, column=7, value="Not Started")
        ws.row_dimensions[i].height = 22

        for col in range(1, 8):
            ws.cell(row=i, column=col).border = thin_border()

        dv_status.add(ws.cell(row=i, column=7).coordinate)
        dv_effort.add(ws.cell(row=i, column=6).coordinate)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    del wb[wb.sheetnames[0]]  # remove default empty sheet

    build_client_info(wb)
    _, section_meta = build_scoring(wb)
    build_dashboard(wb, section_meta)
    build_roadmap(wb)

    output_path = os.path.join(os.getcwd(), OUTPUT_FILE)
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
